# Excel Spreadsheets project

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename) # loading workbook
    sheet = wb['Sheet1'] # loading a sheet from workbook
    cell = sheet['a1'] # accessing the column
    #sheet.cell(1,1) # another way to access the first cell
    print(cell.value)
    print(sheet.max_row) # to get total rows

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        updated_price = cell.value * 0.9
        updated_price_cell = sheet.cell(row, 4)
        updated_price_cell.value = updated_price

    #wb.save('transaction_updated.xlsx')

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart)

    wb.save(filename)

process_workbook('transaction.xlsx')